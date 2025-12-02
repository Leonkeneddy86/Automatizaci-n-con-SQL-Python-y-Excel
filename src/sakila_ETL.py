# src/sakila_ETL.py
import os
from pathlib import Path
import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

#problema ruta absoluta/relativa, ejecutar main.py con .config y ejecutar sakila_ETL con config
from .config import SQLALCHEMY_DATABASE_URI, CSV_PATH, EXCEL_PATH

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def get_engine():
    """Crear engine SQLAlchemy."""
    engine = create_engine(SQLALCHEMY_DATABASE_URI, pool_pre_ping=True)
    return engine


def extract_data(engine):
    """
    Extrae datos relevantes de la base de datos Sakila.
    Devuelve un dataframe con pagos relacionados a rentals/customers/ciudades/paises.
    """
    logging.info("Extrayendo datos de la base de datos...")
    query = """
    SELECT
        p.payment_id,
        p.payment_date,
        p.amount,
        r.rental_id,
        r.rental_date,
        c.customer_id,
        CONCAT(c.first_name, ' ', c.last_name) AS customer_name,
        a.address,
        ci.city,
        co.country
    FROM payment p
    JOIN rental r ON p.rental_id = r.rental_id
    JOIN customer c ON r.customer_id = c.customer_id
    LEFT JOIN address a ON c.address_id = a.address_id
    LEFT JOIN city ci ON a.city_id = ci.city_id
    LEFT JOIN country co ON ci.country_id = co.country_id
    ORDER BY p.payment_date;
    """
    with engine.connect() as conn:
        df = pd.read_sql_query(text(query), conn)
    logging.info(f"Registros extraídos: {len(df)}")
    return df


def transform_data(df):
    """
    Realiza transformaciones y agrega tablas resumen:
      - limpieza de fechas
      - agregados por cliente (total pagado, número de rentals, última fecha)
      - agregados por ciudad y país
    Devuelve: df_raw (pagos), df_customers (agregado por cliente), df_country (resumen por país)
    """
    logging.info("Transformando datos...")
    df = df.copy()
    # Asegurar tipos
    df["payment_date"] = pd.to_datetime(df["payment_date"])
    df["rental_date"] = pd.to_datetime(df["rental_date"])
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)

    # Datos por cliente (agregado)
    df_customers = (
        df.groupby(["customer_id", "customer_name", "city", "country"], dropna=False)
        .agg(
            total_paid=pd.NamedAgg(column="amount", aggfunc="sum"),
            rentals_count=pd.NamedAgg(column="rental_id", aggfunc=lambda x: x.nunique()),
            avg_payment=pd.NamedAgg(column="amount", aggfunc="mean"),
            last_rental=pd.NamedAgg(column="rental_date", aggfunc="max"),
        )
        .reset_index()
    )
    df_customers["avg_payment"] = df_customers["avg_payment"].round(2)
    df_customers["last_rental"] = pd.to_datetime(df_customers["last_rental"]).dt.date

    # Resumen por país y ciudad
    df_country = (
        df.groupby(["country", "city"], dropna=False)
        .agg(total_revenue=pd.NamedAgg(column="amount", aggfunc="sum"),
             rentals_count=pd.NamedAgg(column="rental_id", aggfunc=lambda x: x.nunique()))
        .reset_index()
        .sort_values("total_revenue", ascending=False)
    )

    logging.info("Transformaciones completadas.")
    return df, df_customers, df_country


def save_csv(df_customers, csv_path=CSV_PATH):
    """Guarda dataset pre-procesado (agregado por cliente) en CSV."""
    logging.info(f"Guardando CSV en {csv_path} ...")
    df_customers.to_csv(csv_path, index=False, encoding="utf-8")
    logging.info("CSV guardado.")


def write_excel(df_customers, df_country, df_raw, excel_path=EXCEL_PATH):
    """
    Genera un archivo Excel con 3 hojas:
      - Datos: df_customers (dataset pre-procesado)
      - Tablas dinámicas: df_country (resúmenes por país/ciudad)
      - Dashboard: métricas clave + gráfico (top 10 países por revenue)
    """
    logging.info(f"Generando Excel en {excel_path} ...")
    # Usamos pandas ExcelWriter con engine openpyxl
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_customers.to_excel(writer, sheet_name="Datos", index=False)
        df_country.to_excel(writer, sheet_name="Tablas dinamicas", index=False)
        # Hoja Dashboard: escribimos métricas y la tabla de raw para referencia
        dashboard_df = pd.DataFrame({
            "metric": [
                "fecha_generacion",
                "total_clientes",
                "ingresos_totales",
                "promedio_pago_por_cliente"
            ],
            "value": [
                datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                df_customers["customer_id"].nunique(),
                round(df_raw["amount"].sum(), 2),
                round(df_customers["avg_payment"].mean(), 2)
            ]
        })
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=0)

    # Agregar un gráfico al Dashboard usando openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb["Dashboard"]

    # Preparamos tabla para gráfico: tomamos top 10 países por revenue desde df_country
    top_countries = df_country.groupby("country", as_index=False)["total_revenue"].sum()
    top_countries = top_countries.sort_values("total_revenue", ascending=False).head(10)
    # Escribir la tabla debajo de las métricas
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="country")
    ws.cell(row=start_row, column=2, value="total_revenue")
    for i, row in enumerate(top_countries.itertuples(index=False), start=1):
        ws.cell(row=start_row + i, column=1, value=row.country)
        ws.cell(row=start_row + i, column=2, value=float(row.total_revenue))

    # Crear gráfico de barras
    chart = BarChart()
    chart.title = "Top 10 países por ingresos"
    chart.y_axis.title = "Ingresos"
    chart.x_axis.title = "País"

    data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(top_countries))
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(top_countries))
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.width = 20
    chart.height = 12

    # Insertar gráfico en hoja Dashboard
    ws.add_chart(chart, f"E{start_row}")

    wb.save(excel_path)
    logging.info("Excel generado con éxito.")


def run_etl():
    """Función principal que orquesta todo el pipeline."""
    engine = get_engine()
    df_raw = extract_data(engine)
    df_raw, df_customers, df_country = transform_data(df_raw)
    save_csv(df_customers)
    write_excel(df_customers, df_country, df_raw)
    logging.info("ETL completado. Archivos generados en 'output/' y 'dashboard/'.")